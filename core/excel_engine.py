import os
import re
from datetime import datetime
from typing import List, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils.datetime import from_excel
from openpyxl.styles import Alignment, Font, PatternFill, Border
from openpyxl.styles import Alignment, Font, PatternFill, Border


def safe_filename(name: str) -> str:
    s = str(name).strip()
    s = re.sub(r'[\\/:*?"<>|]', "_", s)
    if not s:
        s = "空值"
    return s[:150]


def make_output_dir(base_dir: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = os.path.join(base_dir, f"output_{ts}")
    os.makedirs(out, exist_ok=True)
    return out


def _unique_xlsx_path(directory: str, base_filename_no_ext: str) -> str:
    base = f"{base_filename_no_ext}.xlsx"
    candidate = os.path.join(directory, base)
    if not os.path.exists(candidate):
        return candidate
    idx = 1
    while True:
        alt = os.path.join(directory, f"{base_filename_no_ext}_{idx}.xlsx")
        if not os.path.exists(alt):
            return alt
        idx += 1


def _fmt_or_general(fmt) -> str:
    if isinstance(fmt, str) and fmt:
        return fmt
    return "General"

DATE_HEADER_KEYS = [
    "时间", "日期", "到货", "到款", "销售时间", "到货时间", "确认时间"
]


def _is_header_date_like(text) -> bool:
    if text is None:
        return False
    s = str(text)
    return any(k in s for k in DATE_HEADER_KEYS)


def _copy_value_and_format(src_cell, dst_cell, src_wb, header_text=None):
    fmt = _fmt_or_general(getattr(src_cell, "number_format", None))
    val = src_cell.value
    try:
        if isinstance(val, (int, float)) and (is_date_format(fmt) or _is_header_date_like(header_text)):
            val = from_excel(val, src_wb.epoch)
            if fmt == "General" and _is_header_date_like(header_text):
                fmt = 'm"月"d"日"'
    except Exception:
        pass
    dst_cell.value = val
    dst_cell.number_format = fmt
    try:
        if src_cell.alignment:
            dst_cell.alignment = src_cell.alignment
        elif _is_header_date_like(header_text):
            dst_cell.alignment = Alignment(horizontal="center")
    except Exception:
        pass
    try:
        dst_cell.font = src_cell.font if src_cell.font else Font()
    except Exception:
        pass
    try:
        dst_cell.fill = src_cell.fill if src_cell.fill else PatternFill()
    except Exception:
        pass
    try:
        dst_cell.border = src_cell.border if src_cell.border else Border()
    except Exception:
        pass
    try:
        dst_cell.alignment = src_cell.alignment if src_cell.alignment else Alignment()
    except Exception:
        pass
    try:
        dst_cell.font = src_cell.font if src_cell.font else Font()
    except Exception:
        pass
    try:
        dst_cell.fill = src_cell.fill if src_cell.fill else PatternFill()
    except Exception:
        pass
    try:
        dst_cell.border = src_cell.border if src_cell.border else Border()
    except Exception:
        pass


def read_headers(file_path: str) -> List[str]:
    try:
        wb = load_workbook(file_path, read_only=False, data_only=True)
    except PermissionError as e:
        raise PermissionError("文件可能被其它程序占用，请关闭后重试") from e
    ws = wb.active
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    wb.close()
    if not row:
        return []
    return [str(c) if c is not None else "" for c in row]


def split_excel(file_path: str, base_output_dir: str, split_column: str, progress_cb=None) -> Tuple[str, List[str]]:
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except PermissionError as e:
        raise PermissionError("源文件被占用或无法读取，请关闭占用程序") from e
    ws = wb.active
    header_cells = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=False), []))
    if not header_cells:
        wb.close()
        raise ValueError("文件为空或无表头")
    header_values = [str(c.value) if c.value is not None else "" for c in header_cells]
    try:
        col_index = header_values.index(split_column)
    except ValueError:
        wb.close()
        raise KeyError(f"未找到列：{split_column}")
    output_dir = make_output_dir(base_output_dir)
    writers = {}
    written_paths = []
    processed_rows = 0
    try:
        for row_cells in ws.iter_rows(min_row=2, values_only=False):
            processed_rows += 1
            key = row_cells[col_index].value
            if key is None:
                if progress_cb and processed_rows % 1000 == 0:
                    progress_cb(processed_rows, len(writers))
                continue
            if key not in writers:
                out_wb = Workbook()
                out_ws = out_wb.active
                for j, c in enumerate(header_cells, start=1):
                    cell = out_ws.cell(row=1, column=j)
                    _copy_value_and_format(c, cell, wb, header_text=c.value)
                fname_base = safe_filename(key)
                out_path = _unique_xlsx_path(output_dir, fname_base)
                writers[key] = (out_wb, out_ws, out_path, 2)
                if progress_cb:
                    progress_cb(processed_rows, len(writers))
            out_wb, out_ws, _path, next_row = writers[key]
            for j, c in enumerate(row_cells, start=1):
                cell = out_ws.cell(row=next_row, column=j)
                _copy_value_and_format(c, cell, wb, header_text=header_cells[j-1].value)
            writers[key] = (out_wb, out_ws, _path, next_row + 1)
            if progress_cb and processed_rows % 1000 == 0:
                progress_cb(processed_rows, len(writers))
    finally:
        wb.close()
        for _, (owb, _ows, path, _next_row) in writers.items():
            owb.save(path)
            owb.close()
            written_paths.append(path)
    return output_dir, written_paths


def merge_excels(file_paths: List[str], base_output_dir: str) -> Tuple[str, str]:
    if not file_paths:
        raise ValueError("未提供可合并的文件")
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    next_row = 1
    wrote_header = False
    for p in file_paths:
        try:
            wb = load_workbook(p, read_only=False, data_only=True)
        except PermissionError as e:
            raise PermissionError("其中一个文件被占用或无法读取，请关闭占用程序") from e
        ws = wb.active
        header_cells = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=False), []))
        if header_cells and not wrote_header:
            for j, c in enumerate(header_cells, start=1):
                cell = merged_ws.cell(row=next_row, column=j)
                _copy_value_and_format(c, cell, wb, header_text=c.value)
            next_row += 1
            wrote_header = True
        for row_cells in ws.iter_rows(min_row=2, values_only=False):
            for j, c in enumerate(row_cells, start=1):
                cell = merged_ws.cell(row=next_row, column=j)
                _copy_value_and_format(c, cell, wb, header_text=header_cells[j-1].value)
            next_row += 1
        wb.close()
    output_dir = make_output_dir(base_output_dir)
    out_path = os.path.join(output_dir, "merged.xlsx")
    merged_wb.save(out_path)
    merged_wb.close()
    return output_dir, out_path